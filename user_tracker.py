import os
import openpyxl
from openpyxl import Workbook
import tkinter as tk
from tkinter import ttk
from datetime import datetime

class UserActionTracker:
    def __init__(self, root):
        self.root = root
        today_str = datetime.today().strftime("%m-%d-%Y")
        self.output_dir = f"User Statistics {today_str}"
        os.makedirs(self.output_dir, exist_ok=True)

        # Fix formatting of the log file path
        self.log_file = os.path.join(self.output_dir, f"user_statistics_{today_str}.xlsx")
        
        # Initialize log file if needed
        self._initialize_log_file()

    def _initialize_log_file(self):
        """Create the Excel log file with headers if it doesn't already exist."""
        if not os.path.exists(self.log_file):
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Actions"
            sheet.append(["Username", "Role", "Action", "Timestamp"])
            wb.save(self.log_file)

    def track_action(self, username, role, action):
        """Record a user action in the Excel log."""
        timestamp = datetime.now().strftime("%m-%d-%Y %H:%M:%S")
        wb = openpyxl.load_workbook(self.log_file)
        sheet = wb.active
        sheet.append([username, role, action, timestamp])
        wb.save(self.log_file)

    def get_action_log(self):
        """Read all logged user actions from the Excel file."""
        log_entries = []
        if os.path.exists(self.log_file):
            wb = openpyxl.load_workbook(self.log_file)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                log_entries.append(row)
        return log_entries

    def display_action_table(self):
        """Display the action log in a table using Tkinter Treeview."""
        top = tk.Toplevel(self.root)
        top.title("User Action Log")
        top.geometry("700x400")

        # Treeview setup
        tree = ttk.Treeview(top, columns=("Username", "Role", "Action", "Timestamp"), show="headings")
        for col in ("Username", "Role", "Action", "Timestamp"):
            tree.heading(col, text=col)
            tree.column(col, width=160)

        # Populate the Treeview
        for entry in self.get_action_log():
            tree.insert("", "end", values=entry)

        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Scrollbar for large logs
        scrollbar = ttk.Scrollbar(top, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")